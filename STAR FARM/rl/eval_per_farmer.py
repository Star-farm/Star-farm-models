"""Per-FARMER evaluation export — structured for cross-farmer comparison.

Runs ONE greedy episode with a trained MAPPO checkpoint in PER-AGENT reward mode and
records, for every farmer and every year: the farm observation (raw, decoded), the action
the policy chose, and the reward (the farmer's own profit) earned that year. Writes an
Excel workbook organised so you can both READ what each number means and COMPARE farmers
side by side:

  * legende            : dictionary of every column / encoding (read this first)
  * synthese_fermiers  : one row per farmer — totals, averages, action mix (the headline)
  * reward_par_an      : matrix year x farmer of the yearly reward (+ GLOBAL), colour-scaled
  * cultivar_par_an    : matrix year x farmer of the cultivar choice (premium highlighted)
  * irrigation_par_an  : matrix year x farmer of the irrigation regime (AWD highlighted)
  * actions_completes  : matrix year x farmer of the full action string
  * detail             : the raw long table (one row per year x farmer)

In per-agent mode rewards[a] IS farmer a's profit delta, so summing over the episode gives
its 25-year profit; the sum across farmers equals the global profit (sanity-checked).

Usage (GAMA server must be running on the port in the config):
    py eval_per_farmer.py [config.yaml] [--best] [--last] [--ckpt PATH] [--out FILE.xlsx]
"""
from __future__ import annotations

import argparse
import os
from collections import defaultdict

import numpy as np
import pandas as pd
import torch
import yaml
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from starfarm_env import StarfarmParallelEnv
from mappo_agent import MAPPOAgent

# Raw rl_obs layout (first 11 dims of the 18-dim observation).
OBS_COLS = ["prev_profit", "fertilizer", "soil_health", "yield_t_ha", "pollution",
            "salinity", "nb_plots", "area_ha", "premium_share", "premium_price",
            "decided_frac"]

LEGEND = [
    ("FEUILLES", ""),
    ("synthese_fermiers", "1 ligne / fermier : profit total et moyen sur l'episode, taille de ferme, etat moyen, et mix d'actions (%). Trie par profit decroissant."),
    ("reward_par_an", "Matrice annee x fermier du profit de l'annee (colonne GLOBAL = somme). Echelle de couleur : rouge=bas, vert=haut."),
    ("cultivar_par_an", "Matrice annee x fermier du cultivar choisi. 'premium' surligne en vert -> montre la diversification premium/standard entre fermes."),
    ("irrigation_par_an", "Matrice annee x fermier du regime d'irrigation. 'AWD' surligne en bleu."),
    ("actions_completes", "Matrice annee x fermier : action complete encodee 'irrigation|qX.XX|pesticide|fertil|cultivar|Ns'."),
    ("detail", "Table brute : 1 ligne par (annee, fermier) avec observation + action + reward."),
    ("", ""),
    ("COLONNES", ""),
    ("reward_year", "Profit realise par le fermier CETTE annee (EUR). C'est la recompense optimisee en mode per-agent."),
    ("total_reward", "Somme des reward_year sur l'episode = profit du fermier sur 25 ans (EUR)."),
    ("mean_annual_reward", "Profit moyen par an (total_reward / nb annees)."),
    ("prev_profit", "Observation : profit de l'annee PRECEDENTE (etat vu au moment de decider)."),
    ("nb_plots / area_ha", "Taille de la ferme : nombre de parcelles / surface (ha). Statiques sur l'episode."),
    ("soil_health", "Sante du sol moyenne (etat agronomique cumulatif)."),
    ("yield_t_ha", "Rendement physique moyen (tonnes / ha)."),
    ("fertilizer", "Quantite moyenne d'engrais appliquee."),
    ("pollution / salinity", "Externalite environnementale locale moyenne (pollution cellule / salinite)."),
    ("premium_share", "Signal marche : part de riz premium dans le paysage (dynamique pendant le tour de decision)."),
    ("premium_price", "Signal marche : prix relatif du cultivar premium (ST25)."),
    ("decided_frac", "Fraction de fermiers ayant deja decide AVANT ce fermier dans l'annee (ordre aleatoire)."),
    ("", ""),
    ("ACTIONS", ""),
    ("irrigation", "CF = inondation continue | AWD = alternance humide/sec."),
    ("irr_qty", "Intensite irrigation [0-1] : CF -> lame d'eau 30-70 mm ; AWD -> seuil de declenchement -50 a -250 mm."),
    ("pesticide", "BAU = conventionnel | IPM = lutte integree."),
    ("fertil", "BAU = conventionnel | durable = fertilisation raisonnee."),
    ("cultivar", "standard (OM5451) | premium (ST25, mieux paye mais sature le marche si trop de monde)."),
    ("seasons", "2 ou 3 saisons de culture par an."),
    ("pct_AWD / pct_IPM / pct_durable / pct_premium / pct_3seasons", "Frequence (%) de chaque option sur l'episode -> profil du fermier."),
]


def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def decode_action(a):
    a = np.clip(np.asarray(a, dtype=float), 0.0, 1.0)
    return dict(irrigation="AWD" if a[0] >= 0.5 else "CF",
               irr_qty=round(float(a[1]), 3),
               pesticide="IPM" if a[2] >= 0.5 else "BAU",
               fertil="durable" if a[3] >= 0.5 else "BAU",
               cultivar="premium" if a[4] >= 0.5 else "standard",
               seasons=3 if a[5] >= 0.5 else 2)


def run_eval(env, agent, order, sequential, rng, stochastic=False):
    """One evaluation episode: greedy by default, or stochastic=True to SAMPLE the policy
    distribution. Returns per_year DataFrame + total_reward dict + sanity global."""
    act_fn = agent.act_sample if stochastic else agent.act_greedy
    year_obs, year_act, year_rew = {}, {}, defaultdict(float)
    total_rew = {a: 0.0 for a in order}
    global_year = defaultdict(float)

    obs, _ = env.reset()
    year = -1
    cur = None
    while True:
        if cur is None or env.needs_new_action:
            year += 1
            cur = {}
            turn = list(order)
            if sequential:
                rng.shuffle(turn)
                for a in turn:
                    o = env.observe_one(a)
                    act = act_fn(o)
                    env.inject_one(a, act)
                    cur[a] = act
                    year_obs[(year, a)] = np.asarray(o, dtype=float)
                    year_act[(year, a)] = np.asarray(act, dtype=float)
                env.end_decision_round()
            else:
                for a in order:
                    act = act_fn(obs[a])
                    cur[a] = act
                    year_obs[(year, a)] = np.asarray(obs[a], dtype=float)
                    year_act[(year, a)] = np.asarray(act, dtype=float)
        nobs, rew, term, trunc, infos = env.step(cur)
        gp = float(next(iter(infos.values())).get("global_profit", 0.0)) if infos else 0.0
        global_year[year] += gp
        for a in order:
            r = float(rew.get(a, 0.0))
            year_rew[(year, a)] += r
            total_rew[a] += r
        obs = nobs
        if (trunc and all(trunc.values())) or (term and any(term.values())):
            break

    n_years = year + 1
    rows = []
    for y in range(n_years):
        for a in order:
            o = year_obs[(y, a)]
            row = {"year": y, "farmer": a, "reward_year": year_rew[(y, a)]}
            for i, col in enumerate(OBS_COLS):
                row[col] = float(o[i])
            row.update(decode_action(year_act[(y, a)]))
            rows.append(row)
    return pd.DataFrame(rows), total_rew, sum(global_year.values()), n_years


def bold_headers(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "B2"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("config", nargs="?",
                    default=os.path.join(os.path.dirname(__file__), "config_mappo3_peragent.yaml"))
    ap.add_argument("--best", action="store_true", help="load best_<filename> (default)")
    ap.add_argument("--last", action="store_true", help="load the last-episode checkpoint instead of best")
    ap.add_argument("--ckpt", default=None, help="explicit checkpoint path")
    ap.add_argument("--out", default=None, help="output .xlsx path")
    args = ap.parse_args()

    cfg_path = args.config
    if not os.path.exists(cfg_path):
        alt = os.path.join(os.path.dirname(os.path.abspath(__file__)), os.path.basename(cfg_path))
        if os.path.exists(alt):
            cfg_path = alt
    cfg = load_config(cfg_path)
    g, t, c = cfg["gama"], cfg["training"], cfg["checkpoint"]

    ckpt_dir = os.path.join(os.path.dirname(os.path.abspath(cfg_path)), c["dir"])
    fname = c["filename"] if args.last else ("best_" + c["filename"])
    ckpt_path = args.ckpt or os.path.join(ckpt_dir, fname)
    out_path = args.out or os.path.join(os.path.dirname(os.path.abspath(cfg_path)),
                                        "eval_per_farmer_peragent.xlsx")

    print(f"Connecting to GAMA at {g['ip_address']}:{g['port']} ...")
    env = StarfarmParallelEnv(
        gaml_experiment_path=g["controler_path"], gaml_experiment_name=g["experiment_name"],
        gama_ip_address=g["ip_address"], gama_port=g["port"])
    env.MIN_DAYS_PER_YEAR = int(g.get("min_days_per_year", env.MIN_DAYS_PER_YEAR))
    env.MAX_DAYS_PER_YEAR = int(g.get("max_days_per_year", env.MAX_DAYS_PER_YEAR))
    env.YEAR_END_MARGIN = int(g.get("year_end_margin", env.YEAR_END_MARGIN))
    env.SUBSTEP_DAYS = int(g.get("substep_days", 0))
    sequential = bool(t.get("sequential_decisions", False))
    env.set_reward_mode(True)  # PER-AGENT
    print(f"Reward mode: PER-AGENT | sequential={sequential} | substep_days={env.SUBSTEP_DAYS}")

    order = sorted(env.possible_agents)
    obs_dim = int(np.prod(env.observation_space(order[0]).shape))
    act_dim = int(np.prod(env.action_space(order[0]).shape))
    if not os.path.exists(ckpt_path):
        raise FileNotFoundError(ckpt_path)
    # Scalability: the actor is per-farmer (shared, no agent id) so it transfers to ANY number
    # of farms. Only the centralized critic depends on n_agents, and it is unused at eval, so
    # we size it to the CHECKPOINT's n_agents to keep loading valid when the env now has more.
    ckpt_n = int(torch.load(ckpt_path, map_location="cpu", weights_only=False).get("n_agents", len(order)))
    agent = MAPPOAgent(
        obs_dim=obs_dim, act_dim=act_dim, n_agents=ckpt_n,
        hidden_dim=t["hidden_dim"], lr=t["lr"], gamma=t["gamma"], gae_lambda=t["gae_lambda"],
        clip_eps=t["clip_eps"], k_epochs=t["k_epochs"], minibatch_size=t["minibatch_size"],
        entropy_coef=t["entropy_coef"], value_coef=t["value_coef"], normalize_obs=t["normalize_obs"],
        critic_hidden_dim=t.get("critic_hidden_dim"))
    agent.load(ckpt_path)
    print(f"Loaded checkpoint: {ckpt_path}")
    print(f"SCALABILITY: env has {len(order)} farmers | policy trained on {ckpt_n} "
          f"(shared actor transfers; critic unused at eval)")

    rng = np.random.default_rng(t.get("seed", 0))
    per_year, total_rew, sanity_global, n_years = run_eval(env, agent, order, sequential, rng)
    env.close()

    # ---- per-farmer synthesis (sorted by total reward, best first) ----
    summ = []
    for a in order:
        sub = per_year[per_year.farmer == a]
        summ.append({
            "farmer": a,
            "total_reward": total_rew[a],
            "mean_annual_reward": total_rew[a] / n_years,
            "nb_plots": float(sub.nb_plots.iloc[0]),
            "area_ha": float(sub.area_ha.iloc[0]),
            "mean_soil_health": sub.soil_health.mean(),
            "mean_yield_t_ha": sub.yield_t_ha.mean(),
            "mean_fertilizer": sub.fertilizer.mean(),
            "mean_pollution": sub.pollution.mean(),
            "mean_salinity": sub.salinity.mean(),
            "pct_AWD": 100.0 * (sub.irrigation == "AWD").mean(),
            "pct_IPM": 100.0 * (sub.pesticide == "IPM").mean(),
            "pct_durable": 100.0 * (sub.fertil == "durable").mean(),
            "pct_premium": 100.0 * (sub.cultivar == "premium").mean(),
            "pct_3seasons": 100.0 * (sub.seasons == 3).mean(),
            "mean_irr_qty": sub.irr_qty.mean(),
        })
    per_farmer = pd.DataFrame(summ).sort_values("total_reward", ascending=False).reset_index(drop=True)
    ranked = list(per_farmer.farmer)  # best farmer first -> column order for all pivots

    # ---- pivots (year x farmer), columns ordered best->worst ----
    def pivot(col):
        return per_year.pivot(index="year", columns="farmer", values=col)[ranked]

    rew_p = pivot("reward_year")
    rew_p["GLOBAL"] = rew_p.sum(axis=1)
    cul_p = pivot("cultivar")
    irr_p = pivot("irrigation")
    per_year["action_str"] = per_year.apply(
        lambda r: f"{r.irrigation}|q{r.irr_qty:.2f}|{r.pesticide}|{r.fertil}|{r.cultivar}|{int(r.seasons)}s", axis=1)
    act_p = pivot("action_str")
    legend = pd.DataFrame(LEGEND, columns=["element", "signification"])

    nf = len(ranked)
    last_farmer = get_column_letter(1 + nf)   # col A = index, farmers B..(1+nf)
    data_rng = f"B2:{last_farmer}{n_years + 1}"

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        legend.to_excel(xl, sheet_name="legende", index=False)
        per_farmer.to_excel(xl, sheet_name="synthese_fermiers", index=False)
        rew_p.to_excel(xl, sheet_name="reward_par_an")
        cul_p.to_excel(xl, sheet_name="cultivar_par_an")
        irr_p.to_excel(xl, sheet_name="irrigation_par_an")
        act_p.to_excel(xl, sheet_name="actions_completes")
        per_year.drop(columns=["action_str"]).to_excel(xl, sheet_name="detail", index=False)

        sh = xl.sheets
        for name in ("synthese_fermiers", "reward_par_an", "cultivar_par_an",
                     "irrigation_par_an", "actions_completes", "detail"):
            bold_headers(sh[name])
        # colour scale on the per-year reward matrix (farmer columns only, excl GLOBAL)
        sh["reward_par_an"].conditional_formatting.add(data_rng, ColorScaleRule(
            start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50,
            mid_color="FFEB84", end_type="max", end_color="63BE7B"))
        # highlight premium / AWD so diversification is visible at a glance
        sh["cultivar_par_an"].conditional_formatting.add(data_rng, CellIsRule(
            operator="equal", formula=['"premium"'], fill=PatternFill("solid", fgColor="C6EFCE")))
        sh["irrigation_par_an"].conditional_formatting.add(data_rng, CellIsRule(
            operator="equal", formula=['"AWD"'], fill=PatternFill("solid", fgColor="BDD7EE")))

    gt = sum(total_rew.values())
    print(f"\nEpisode: {n_years} years | sum(farmers)={gt:,.0f} | sanity sum(infos global)={sanity_global:,.0f}")
    print("\nPer-farmer 25-year profit (best first):")
    for _, r in per_farmer.iterrows():
        print(f"  {r.farmer:>16s} total={r.total_reward:>12,.0f}  premium={r.pct_premium:>4.0f}%  "
              f"AWD={r.pct_AWD:>4.0f}%  IPM={r.pct_IPM:>4.0f}%  durable={r.pct_durable:>4.0f}%  "
              f"soil={r.mean_soil_health:>5.2f}  yield={r.mean_yield_t_ha:>5.2f}")
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
